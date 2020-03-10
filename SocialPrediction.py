import openpyxl
import pandas as pd
import xlrd
#import Workbook
from nltk.util import ngrams
from nltk.corpus import treebank
from openpyxl import load_workbook
import re
import nltk
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
from nltk.probability import FreqDist
#from pycontractions import Contractions
import matplotlib.pyplot as plt
from nltk.stem.wordnet import WordNetLemmatizer

wb= load_workbook(filename = 'retarded data.xlsx')
sheet=wb.active
wb1 = openpyxl.Workbook()
sheet1 = wb1.active
WNlemma = nltk.WordNetLemmatizer()
sheet1.cell(row=1, column=1).value= "name"
sheet1.cell(row=1, column=2).value= "text"
for i in range(2,sheet.max_row+1):
    sheet1.cell(row=i,column=1).value = sheet.cell(row=i,column=1).value
    if (sheet.cell(row=i, column=2).value == None):
        continue
    j=2
    k=2
    while(1):
        if(sheet.cell(row=i , column=j).value== None):
            break
        post = sheet.cell(row=i, column=j).value
        post = re.sub(r"[^A-Za-z ]", " ", post)
        post=post.lower()
        print(post)
        tokenized_word=word_tokenize(post)
        print(tokenized_word)
        postageboi=[postageboi for postageboi in tokenized_word if postageboi.isalpha()]
        refined_list = [WNlemma.lemmatize(t, pos='v') for t in postageboi]
        stop_words=set(stopwords.words("english"))
        filtered_sent=[]
        for w in refined_list:
            if w not in stop_words:
                filtered_sent.append(w)
        fdist = FreqDist(filtered_sent)
        print(filtered_sent)
        print("\n\n\n\n\n\n")
        if (j==2):
            stringnotsting = ' '.join(filtered_sent)
        else:
            print(stringnotsting)
            stringnotsting2 = ' '.join(filtered_sent)
            stringnotsting = stringnotsting + " " + stringnotsting2
            sheet1.cell(row=i,column=2).value = stringnotsting

        j=j+1
wb1.save("D:\\AI\\SocialPrediction\\demo.xlsx")

df1 = pd.read_excel("demo.xlsx")
total_rows1=len(df1.axes[0])
filtered_df1 = df1[df1['text'].notnull()]
print(filtered_df1)
print(len(filtered_df1))
export_excel = filtered_df1.to_excel (r'D:\AI\SocialPrediction\demo.xlsx', index = None, header=True) #Don't forget to add '.xlsx' at the end of the path
