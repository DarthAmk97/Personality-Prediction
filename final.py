import openpyxl
import pandas as pd
from openpyxl import load_workbook
import re
import nltk
# nltk.data.path.append("C:\\nltk")
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
from nltk.probability import FreqDist

#Data Pre Processing; Lemmatize, Regular Expression, Stop Words, tokenization, removal of null rows
wb= load_workbook(filename = 'retarded data.xlsx')
sheet=wb.active
wb1 = openpyxl.Workbook()
sheet1 = wb1.active
WNlemma = nltk.WordNetLemmatizer()
sheet1.cell(row=1, column=1).value= "name"
sheet1.cell(row=1, column=2).value= "text"
for i in range(2,sheet.max_row+1):
    sheet1.cell(row=i,column=1).value = sheet.cell(row=i,column=1).value
    if (sheet.cell(row=i, column=2).value == None):
        continue
    j=2
    k=2
    while(1):
        if(sheet.cell(row=i , column=j).value== None):
            break
        post = sheet.cell(row=i, column=j).value
        post = re.sub(r"[^A-Za-z ]", " ", post)
        post=post.lower()
        tokenized_word=word_tokenize(post)
        postageboi=[postageboi for postageboi in tokenized_word if postageboi.isalpha()]
        print(postageboi)
        refined_list = [WNlemma.lemmatize(t, pos='v') for t in postageboi]
        stop_words=set(stopwords.words("english"))
        filtered_sent=[]
        for w in refined_list:
            if w not in stop_words:
                filtered_sent.append(w)
        fdist = FreqDist(filtered_sent)
        print(filtered_sent)
        if (j==2):
            stringnotsting = ' '.join(filtered_sent)
        else:
            # print(stringnotsting)
            stringnotsting2 = ' '.join(filtered_sent)
            stringnotsting = stringnotsting + " " + stringnotsting2
            sheet1.cell(row=i,column=2).value = stringnotsting

        j=j+1
wb1.save("demo.xlsx")

df1 = pd.read_excel("demo.xlsx")
total_rows1=len(df1.axes[0])
filtered_df1 = df1[df1['text'].notnull()]
# print(filtered_df1)
# print(len(filtered_df1))
export_excel = filtered_df1.to_excel (r'demo.xlsx', index = None, header=True)




#TF-IDF to achieve term frequncies, document frequencies, and then to increase count in emotion lexicon excel file to lean towards a certain set of emotion


import math
from openpyxl import load_workbook

map = {}
smap = {}

df = {}
idf = {}
words = list()
s = set()

filepath = "demo.xlsx"
wb = load_workbook(filepath)
sheet = wb.active

N = sheet.max_row - 1

for i in range(2, sheet.max_row + 1):
    line = str(sheet.cell(row=i, column=2).value)
    words = line.split()
    for j in words:
        s.add(j)
        if j not in map:
            map[j] = {}
            map[j] = {i: 1}
        else:
            if i not in map[j]:
                map[j].update({i: 0})
            freq = map[j][i]
            freq = freq + 1
            map[j][i] = freq
for i in s:
    keys = map[i].keys()
    df[i] = keys.__len__()

for i in s:
    key = df[i]
    idf[i] = math.log10(N / key)

for i in s:
    keys = map[i].keys()
    for j in keys:
        freq = map[i][j]
        tf = 1 + math.log10(freq)
        map[i][j] = tf

for i in s:
    keys = map[i].keys()
    for j in keys:
        d = idf[i]
        t = map[i][j]
        map[i][j] = t * d;

# filepath="demo.xlsx"
# wb=load_workbook(filepath)
# sheet=wb.active

filepath="Emotion_Lexicon.xlsx"
wb2 = load_workbook(filepath)
sheet2 = wb2.active

filepath="tf-idf.xlsx"
wb3 = load_workbook(filepath)
sheet3 = wb3.active

# arr = ["anger" , "anticipation" , "disgust" , "fear" , "joy" , "negative" , "positive" , "sadness" , "surprise" , "trust" , "Charged"]
lex = {}

for i in range(2,sheet2.max_row + 1):
   key = sheet2.cell(row=i, column=1).value
   lex[key] = {}
   j = 2
   list = []
   while(1):
       if (sheet2.cell(row=i, column=j).value == None):
           break
       if(sheet2.cell(row=i, column=j).value == 1):
           list.append(j-2)
       j = j + 1
   lex[key] = list

for i in range(2,sheet.max_row + 1):
   for j in range(1,12  ):
       sheet3.cell(row=i , column=j).value = 0

data = {}
j=2
k=0
for i in range(2,sheet.max_row + 1):
   if(sheet.cell(row=i,column=j).value == None):
       continue

   data[i] = {}
   sen = sheet.cell(row=i , column=j).value
   words = sen.split()
   for word in words:
       if word in lex:
           types = lex[word]
           # print(types)
           for k in types:
               val = sheet3.cell(row=i, column= k+1).value
               val = val + map[word][i]
               sheet3.cell(row=i, column=k + 1).value = val

wb3.save("tf-idf.xlsx")



import numpy as np
from openpyxl import load_workbook
import pandas as pd

df = pd.read_excel("tf-idf.xlsx")

pers = { 0: [ "joy","positive","trust","charged"] ,
        1: [ "joy","positive","trust","surprised"],
        2: [ "joy","positive","surprised","charged"] ,
        3: [ "disgust","fear","negative", "anticipation"] ,
        4: [ "anger","anticipation","fear","negative","sadness"]
    }
per = ["Extraversion", "Agreeableness", "Openness", "Conscientiousness", "Neuroticism"]
arank = df.apply(np.argsort, axis=1)
ranked_cols = df.columns.to_series()[arank.values[:,::-1][:,:3]]
# print(ranked_cols)
new_frame = pd.DataFrame(ranked_cols, index=df.index)
# print(new_frame)

from pandas import ExcelWriter
writer = ExcelWriter('pers.xlsx')
new_frame.to_excel(writer,'Sheet1',index=False)
writer.save()

filepath="pers.xlsx"
wb = load_workbook(filepath)
sheet = wb.active

filepath="demo.xlsx"
wb1=load_workbook(filepath)
sheet1=wb1.active

lex = {}
for i in range(2,sheet.max_row + 1):
   lex[i] = {}
   j = 1
   list = []
   for j in range(1,sheet.max_column+1):
       list.append(sheet.cell(row=i, column=j).value)
   lex[i] = list
# print(lex)

count = [0,0,0,0,0]
for i,val in lex.items():
    count = [0, 0, 0, 0, 0]
    for j in val:
        for k in range(0,5):
            if j in pers[k]:
                count[k] = count[k]+1
    # print(count)
    p = np.argsort(count)
    # print(p)
    # print(i)
    # sheet1.cell(row=i, column=3).value = per[p[3]]
    sheet1.cell(row=i, column=3).value = per[p[4]]
# wb.save("pers.xlsx")
wb1.save("demo.xlsx")






from sklearn import metrics
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.model_selection import train_test_split
from openpyxl import load_workbook

wb= load_workbook(filename = 'demo.xlsx')
sheet=wb.active
j=2
x = [0 * 91]
y = [0 * 91]
y[0]=sheet.cell(row=2, column=3).value
for i in range(3,sheet.max_row+1):
    if (sheet.cell(row=i, column=3).value == None):
        continue
    y.append( sheet.cell(row=i, column=3).value )
cv = CountVectorizer()
i=3
x[0]=sheet.cell(row=2, column=2).value
while(1):
    if(i>sheet.max_row):
        break
    if (sheet.cell(row=i, column=2).value == None):
        i=i+1
        continue
    x.append( sheet.cell(row=i, column=2).value )
    i=i+1
x = cv.fit_transform(x).toarray()
X_train, X_test, y_train, y_test = train_test_split(x,y,test_size = 0.2, random_state = 21)

from sklearn.naive_bayes import GaussianNB
from sklearn.metrics import confusion_matrix

classifier = GaussianNB();
classifier.fit(X_train, y_train)

y_pred = classifier.predict(X_test)

cm = confusion_matrix(y_test, y_pred)

acc = metrics.accuracy_score(y_test,y_pred)
print("Our System is",acc*100,"% accurate")

