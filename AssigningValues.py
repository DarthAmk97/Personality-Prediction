#fix column beginning from 2

from openpyxl import load_workbook

filepath="testingtheory123.xlsx"
wb=load_workbook(filepath)
sheet=wb.active

filepath="Emotion_Lexicon.xlsx"
wb2 = load_workbook(filepath)
sheet2 = wb2.active

filepath="tf-idf.xlsx"
wb3 = load_workbook(filepath)
sheet3 = wb3.active

# arr = ["anger" , "anticipation" , "disgust" , "fear" , "joy" , "negative" , "positive" , "sadness" , "surprise" , "trust" , "Charged"]
lex = {}

for i in range(2,sheet2.max_row + 1):
   key = sheet2.cell(row=i, column=1).value
   lex[key] = {}
   j = 2
   list = []
   while(1):
       if (sheet2.cell(row=i, column=j).value == None):
           break
       if(sheet2.cell(row=i, column=j).value == 1):
           list.append(j-2)
       j = j + 1
   lex[key] = list
print(lex)

for i in range(2,sheet.max_row + 1):
   for j in range(2,13):
       sheet3.cell(row=i , column=j).value = 0

data = {}
j=2
k=0
# count = 0
for i in range(2,sheet.max_row + 1):
   if(sheet.cell(row=i,column=j).value == None):
       continue

   data[i] = {}
   # print("i: " + str(i))
   sen = sheet.cell(row=i , column=j).value
   words = sen.split()
   for word in words:
       if word in lex:
           # count = count +1
           types = lex[word]
           for k in types:
               val = sheet3.cell(row=i, column= k+2).value
               val = val + 1
               sheet3.cell(row=i, column=k + 2).value = val
               # print(sheet3.cell(row=i, column=k + 2).value)

wb3.save("tf-idf.xlsx")
#print(count)

