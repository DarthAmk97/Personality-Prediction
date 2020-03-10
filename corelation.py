import numpy as np
from openpyxl import load_workbook
import pandas as pd

df = pd.read_excel("tf-idf.xlsx")
pers = { 0: [ "joy","positive","trust","charged"] ,
        1: [ "joy","positive","trust","surprised"],
        2: [ "joy","positive","surprised","charged"] ,
        3: [ "disgust","fear","negative","positive, anticipation"] ,
        4: [ "anger","anticipation","fear","negative","sadness"]
    }
per = ["Extraversion", "Agreeableness", "Openness", "Conscientiousness", "Neuroticism"]
arank = df.apply(np.argsort, axis=1)
ranked_cols = df.columns.to_series()[arank.values[:,::-2][:,:2]]
print(ranked_cols)
new_frame = pd.DataFrame(ranked_cols, index=df.index)
print(new_frame)
# print(pers)
#
#
from pandas import ExcelWriter
#
#
writer = ExcelWriter('pers.xlsx')
new_frame.to_excel(writer,'Sheet1',index=False)
writer.save()

filepath="pers.xlsx"
wb = load_workbook(filepath)
sheet = wb.active

lex = {}
for i in range(2,sheet.max_row + 1):
   lex[i] = {}
   j = 1
   list = []
   for j in range(1,sheet.max_column+1):
       list.append(sheet.cell(row=i, column=j).value)
   lex[i] = list
print(lex)

count = [0,0,0,0,0]
for i,val in lex.items():
    count = [0, 0, 0, 0, 0]
    for j in val:
        for k in range(0,5):
            if j in pers[k]:
                count[k] = count[k]+1
    print(count)
    p = np.argsort(count)
    print(p)
    print(i)
    sheet.cell(row=i, column=3).value = per[p[3]]
    sheet.cell(row=i, column=4).value = per[p[4]]
wb.save("D:\\AI\\SocialPrediction\\pers.xlsx")
