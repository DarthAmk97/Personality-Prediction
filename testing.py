import pandas as pd
import re
import nltk
from nltk.corpus import stopwords
from nltk.stem.porter import PorterStemmer
import  sklearn
from sklearn import metrics
from sklearn.feature_extraction.text import CountVectorizer
import pandas as pd
from sklearn import datasets, linear_model
from sklearn.model_selection import train_test_split
from matplotlib import pyplot as plt
from openpyxl import load_workbook

wb= load_workbook(filename = 'testingtheory123.xlsx')
sheet=wb.active
j=2
x = [0 * 91]
y = [0 * 91]
y[0]=sheet.cell(row=2, column=3).value
for i in range(3,sheet.max_row+1):
    if (sheet.cell(row=i, column=3).value == None):
        continue
    y.append( sheet.cell(row=i, column=3).value )
# print(y)
cv = CountVectorizer()
# for i in range(2,sheet.max_row+1):
i=3
x[0]=sheet.cell(row=2, column=2).value
while(1):
    if(i>sheet.max_row):
        break
    if (sheet.cell(row=i, column=2).value == None):
        i=i+1
        continue
    x.append( sheet.cell(row=i, column=2).value )
    i=i+1
# print(x)
x = cv.fit_transform(x).toarray()


# splitting the data set into training set and test set
# from sklearn.cross_validation import train_test_split
# import sklearn.cross_validation
# from sklearn.cross_validation import train_test_split
X_train, X_test, y_train, y_test = train_test_split(x,y,test_size = 0.3, random_state = 1)

# X_train, X_test, y_train, y_test = train_test_split(
# 		X, y, test_size = 0.25, random_state = 0)

# fitting naive bayes to the training set
from sklearn.naive_bayes import GaussianNB
from sklearn.metrics import confusion_matrix

classifier = GaussianNB();
classifier.fit(X_train, y_train)

# predicting test set results
y_pred = classifier.predict(X_test)
# print(y_pred)

# making the confusion matrix
cm = confusion_matrix(y_test, y_pred)

acc = metrics.accuracy_score(y_test,y_pred)
print("Our System is",acc*100,"% accurate")